import sys
import openpyxl as px #Extracts data from Excel files

if len(sys.argv) < 4:
    print ("Use: <program file name> <input_file_name.xlsx> <input_file_name.txt> <output_file_name.txt>")
    sys.exit(0)

#Creates dictionary
filein1 = sys.argv[1]
filein1_wb = px.load_workbook(filein1)
filein1_sheet = filein1_wb.active

list1 = []
for i in range(1,filein1_sheet.max_row + 1):
    list1.append((filein1_sheet.cell(row=i,column=1).value,filein1_sheet.cell(row=i,column=2).value))

filein1_dict = dict(list1) #Dictionary of a sequence of nine nucleotides and its scores


filein = open(sys.argv[2], 'r')
fin = filein.read()

fout = open(sys.argv[3], 'w')

# split sequences by '>'
lista1 = str.split(fin, '>')[1:] #the file header is discarded

def evaluation(seq):
	"Receives a sequence of three codons and search in the dictionary its score"
	return filein1_dict.get(seq)


def main():
    
    for element in lista1:
        lista2 = str.replace(element, '\n', '')
        lista3 = str.split(lista2, '\r')

        names = []
        sequences = []
        try:
            for element in lista3:
                e = element.split('ATG',1)
                names.append(e[0])
                sequences.append(e[1])
        except IndexError:
            pass
        
        for i in sequences:
            seq = str(i[3:12])
            n = evaluation(seq)
            e = str(names)
            e1 = e.replace("['",'')
            e2 = e1.replace("']",'')
            fout.write('> ' + e2 +' '+ str(seq) +" "+str(n) + '\n')
          	


    print "Done!"
    filein.close()

main()


